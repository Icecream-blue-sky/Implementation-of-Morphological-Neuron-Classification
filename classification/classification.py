# -*- coding: utf-8 -*-
import numpy as np
import os
import re
import sys
import copy
import time
import pandas as pd
import _pickle as cPickle
from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC
from sklearn.feature_selection import RFECV
from sklearn.model_selection import train_test_split
from sklearn import metrics
import matplotlib.pyplot as plt
dir_bt = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, dir_bt)
import btmorph
#load swc file and return stats of swctree
def loaddata(path):
    swc_tree = btmorph.STree2()
    swc_tree.read_SWC_tree_from_file(path)
    stats = btmorph.BTStats(swc_tree)
    return swc_tree,stats

#get max depth of the dataset using STree2.max_height
#顺便绘制箱线图，观察每个类别神经元depth的分布
#一种优化思路是：在这个方法里用list存储swc_stats对象，decompose时直接从list中提取，就不用重复load
#todo:1、需要在这里把没有soma的不完整swc文件找出来，并及时去除，否则影响极大。这一部分目前只是找出来，但没有去除
#2、对multiple cylinder soma类型的swc文件分解不正确，需要重新分解，好像btmorph已经考虑了这种类型
def get_maxlevel(root_path,class_num):
    start_time = time.time()
    print('finding the max_depth an min_depth of the dataset...')
    #C1-C(class_num)
    label_list = []
    for i in range(1,class_num+1):
        label_list.append('C'+str(i))
    #存储所有神经元的depth信息
    #每行存储一类神经元depth信息
    data = {}
    max_level = 0
    min_level = 10000
    i = 1
    for class_index in range(1,class_num+1):
        data[label_list[class_index-1]] = []
        folder_path = os.path.join(root_path,str(class_index),'rawdata')
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path,file_name)
            swc_tree,swc_stats = loaddata(file_path)
            neuron_depth = swc_tree.max_height - 1
            data[label_list[class_index-1]].append(neuron_depth)
            #find max_depth of the dataset
            if neuron_depth > max_level:
                max_level = neuron_depth
            if neuron_depth < min_level:
                min_level = neuron_depth
            print('the %d th file %s has been loaded!'%(i,file_path))
            #去除异常值，将depth大于35的神经元进行替换
            if neuron_depth > 35:
                print('file %s have to be removed!'%file_path)
                os.remove(file_path)
            i = i+1
    print('the max_depth of the dataset is %d, the min_depth of the dataset is %d'%(max_level,min_level))
    #绘制depth箱线图
    df = pd.DataFrame(data)
    df.plot.box()
    plt.grid(linestyle='--',alpha=0.3)
    save_path = os.path.join(root_path,'depth_distribution.png')
    plt.savefig(save_path)
    print('the result has been saved to %s'%save_path)
    plt.show()
    print('total time: ',time.time()-start_time)
    return max_level


#decompose by depth
#result store in 'save_folder+"/"+str(depth_level)+"/"+neuron_name+"_"+str(level)+'.swc''
def decompose_depth(swc_stats,neuron_name,depth_level,save_folder):
    save_folder = os.path.join(save_folder,'decompose_by_depth')
    swc_stats.write_subtree_by_level(depth_level+1,save_folder,neuron_name)

#decompose by height
#assume that every neuron can reach max_level
#result store in 'save_folder+"/"+str(height_level)+"/"+neuron_name+"_"+str(height_level)+'.swc''
def decompose_height(swc_stats,neuron_name,neuron_depth,max_level,height_level,save_folder):
    save_folder = os.path.join(save_folder,'decompose_by_height')
    #turn height to depth
    depth_level = max_level - height_level
    #if the branch at depth_level exsits
    if depth_level <= neuron_depth:
        swc_stats.write_subtree_by_range(depth_level,neuron_depth+1,save_folder,neuron_name,height_level)
    else:
        #empty folder
        #neuron doesn't have dendritic at this height_level
        save_path = save_folder + "/" + str(height_level) + "/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)

#decompose by slice
#result store in 'save_folder+neuron_name+'_'+"/slice_"+str(slice_depth)+'_'+str(slice_size)+"/" + neuron_name+"_"+"str(slice_depth)+"_"+str(slice_size)+'.swc'
#only consider slice depth:0-10 and slice size:1-20 for temporary
def decompose_slice(swc_stats,neuron_name,neuron_depth,slice_depth,slice_size,save_folder):
    save_folder = os.path.join(save_folder,'decompose_by_slice')
    #todo:记得改回来
    if slice_depth <= neuron_depth:
        swc_stats.write_subtree_by_slice_range(slice_depth,slice_depth+slice_size,save_folder,neuron_name)

#若swc_stats结构太复杂，使用cPickle会堆栈溢出；则直接重新load
#type标记采取上述哪个策略:0,第一种；1，第二种
def get_swc_stats(file_path,type,swc_stats_string):
    if type:
        swc_tree,swc_stats = loaddata(file_path)
        return swc_stats
    return cPickle.loads(swc_stats_string)

#class_num:num of neuron classes
#max_level:the max depth to be decomposed
#since after each decompose the swc_stats(and swc_tree) will be changed,so for each decompose,only transfer deepcopy of swc_stats
#use deepcopy of swc_stas to avoid changing swc_tree
#优化思路：考虑将swc_tree和swc_stats写入本地文件暂存，需要时读取出来，每次都是一个新的对象，替代deepcopy方法和每次重新loaddata
def decompose(root_path,class_num,max_level,slice_max_depth,slice_max_size):
    #我的电脑的最大值，不同电脑不一样
    sys.setrecursionlimit(12750)
    start_time = time.time()
    print('start decompose...')
    max_depth = 0
    i = 1
    #1到4的slice要重新计算，不要忘记了！
    for class_index in range(1,class_num+1):
        folder_path = os.path.join(root_path,str(class_index),'rawdata')
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path,file_name)
            swc_tree,swc_stats = loaddata(file_path)
            type = 0
            #save the object swc_stats to avoid be changed later
            #catch the error:maximum recursion depth exceeded while pickling an object
            try:
                swc_stats_string = cPickle.dumps(swc_stats)
            except RuntimeError as e:
                print(e)
                swc_stats_string = ''
                type = 1
            #neuron_depth = max_level of neuron - 1
            neuron_depth = swc_tree.max_height - 1
            #find max_depth of the dataset
            if neuron_depth > max_depth:
                max_depth = neuron_depth
            neuron_name = re.sub('(.CNG.swc)','',file_name)
            save_folder = os.path.join(root_path,str(class_index))
            for level in range(0,max_level+1):
                decompose_depth(get_swc_stats(file_path,type,swc_stats_string),neuron_name,level,save_folder)
                #after each decompose swc_stats will be changed, so reload the tree.(have tried copy.deepcopy, occur error)
                decompose_height(get_swc_stats(file_path,type,swc_stats_string),neuron_name,neuron_depth,max_level,level,save_folder)
            for slice_depth in range(0,slice_max_depth+1):
                for slice_size in range(1,slice_max_size+1):
                    decompose_slice(get_swc_stats(file_path,type,swc_stats_string),neuron_name,neuron_depth,slice_depth,slice_size,save_folder)
            print('the %d th file %s has been decomposed!'%(i,file_path))
            i = i+1
    print('end decompose!')
    print('the max_depth of the dataset is %d'%max_depth)
    print('Total decompose time：',time.time() - start_time,' seconds')

#deal with the measure data from L-Measure .xls.
#in this step, I will create a output.xlsx for each classify task like \rawdata\....swc and \decompose_by_depth\1\...swc.save locally.
#then in classify step, directly read the output.xlsx.
#这个函数只要将decompose和原始neuron的L-Measure测量数据做一个归类，生成output.xlsx，放到指定文件夹，方便后续分类
#记得增加‘measure_result'目录，存储所有分类好的output.xlsx
#优化思路：减少文件IO操作，对每个output.xlsx文件只读写一次。方法是，初始化一个output_dir数组，保存每个output.xlsx的上级目录名，且元素不可重复。
#再初始化一个output二维数组，其每一行对应一个output.xlsx应该写入的neuron数据，每读取一个neuron的特征信息，就加入到output数组相应的位置。
#output_dir和output全部读写完后，最后再统一写入相应的output.xlsx文件。
#优化版
#再究极优化的思路：因为现在情况是，数据集太大，全放入一个excel文件的一个sheet不现实，暂时分成了10个excel，但如果把10个excel合并到一个超级大的excel中，分配到不同sheet
#然后一次性读取10个class的所有的neuron数据，一次性读取
#或者轮流读取每个excel，进行dataFrame合并，生成一个大的DataFrame
def preprocess_data2(path,save_folder,class_num):
    start_time = time.time()
    #记得改回来
    save_folder = os.path.join(save_folder,'measure_result')
    print('start to preprocess...')
    print('loading data from %s...'%path)
    data = pd.DataFrame()
    #这一步是一次性load所有excel到一个dataframe，一定要处理excel最后一行是空行的情况，不要把空行也读取进来。
    #方法是DataFrame.dropna
    for filename in os.listdir(path):
        path2 = os.path.join(path,filename)
        data = data.append(pd.read_excel(path2,header=None),ignore_index=True)
        print('have loaded data from %s !'%path2)
    data = data.dropna()
    data = data.reset_index(drop=True)
    #get total rows of the data
    nums = data.shape[0]
    #存储每个output.xlsx对应的neuron_data以及存储路径
    #key:存储路径,value:所有对应的neuron-data,type:DataFrame
    output_dict = {}
    running_progress = 0.0
    #打印进度的辅助字典
    state = {}
    for i in range(10,101,10):
        state[i] = 0
    for i in range(0,nums,22):
        neuron_info = data.loc[i,0]
        neuron_info = neuron_info.split('\\')
        #parse neuron_info to find which folder it should be saved
        #and write it into output.xlsx
        for j in range(0,len(neuron_info)):
            if neuron_info[j] == 'train_dataset':
                #get the 'rawdata' or 'decompose_by_depth' or 'decompose_by_height' or 'decompose_by_slice'
                if neuron_info[j+2] == 'rawdata':
                    #the save path is savefolder/rawdata/output.xlsx
                    #todo: 改为output1.xlsx，记得改回来
                    save_path = os.path.join(save_folder,neuron_info[j+2],'output.xlsx')
                elif neuron_info[j+2] in ['decompose_by_depth','decompose_by_height','decompose_by_slice']:
                     save_path = os.path.join(save_folder,neuron_info[j+2],neuron_info[j+3],'output.xlsx')
        #get 22 feature data of a neuron
        #type Pandas.DataFrame
        neuron_data = data.loc[range(i,i+22)]
        if not save_path in output_dict:
            output_dict[save_path] = neuron_data
        else:
            output_dict[save_path] = output_dict[save_path].append(neuron_data,ignore_index=True)
        #添加一个显示运行进度的程序，每10%显示一次已经运行进度
        running_progress = (i+1)/nums
        if int(running_progress*100) in state:
            if not state[int(running_progress*100)]:
                print('running_progress %f %%'%(running_progress*100))
                state[int(running_progress*100)] = 1
    #start to save
    from openpyxl import load_workbook
    for save_path in output_dict.keys():
        #save the neuron to new output.xlsx
        if not os.path.exists(save_path):
            os.makedirs(os.path.dirname(save_path))
            output_dict[save_path].to_excel(save_path,header=None,index=None)
            print('file %s has been preprocessed!'%save_path)
            continue
        #print('file %s has already been preprocessed, pass!'%save_path)
        book = load_workbook(save_path)
        writer = pd.ExcelWriter(save_path,engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title,ws) for ws in book.worksheets)
        sheet = book.worksheets[0]
        output_dict[save_path].to_excel(writer,sheet_name=sheet.title,header=None,index=None,startrow=sheet.max_row)
        book.save(save_path)
        writer.save()
        writer.close()
        print('file %s has been preprocessed!'%save_path)
    print('fininsh preprocess!')
    print('Total preprocess time: ',time.time() - start_time,' seconds')

#for real classify
#data_folder: /train_dataset/measure_data
#有一个潜在的疑问：在进行rfecv降维时，是用完整数据降维一次，选出一些代表性特征，还是每次分类都降维？
#k_fold:StratifiedKFold(k_fold)
#test_size:测试集占总数据集的比例
#todo:优化思路：可以思考下如何在分类时自动去除神经元数量少于k_fold的神经元类别或者将其打出来，手工去除，再或者根据不同情况动态调节k_fold
def classify(data_folder,class_nums,k_fold,test_size):
    start_time = time.time()
    print('start to classify...')
    data_folder = os.path.join(data_folder,'measure_result')
    dir_name = os.listdir(data_folder)
    #path of all output.xlsx
    path_list = []
    #存储有问题的数据位置，人工更改
    problem_pathlist = []
    for i in dir_name:
        if i == 'rawdata':
            path_list.append(os.path.join(data_folder,i,'output.xlsx'))
        elif i in ['decompose_by_depth','decompose_by_height','decompose_by_slice']:
            for j in os.listdir(os.path.join(data_folder,i)):
                path_list.append(os.path.join(data_folder,i,j,'output.xlsx'))
    for path in path_list:
        save_path = os.path.join(os.path.dirname(path),'classify_result.xlsx')
        if os.path.exists(save_path):
            print('file %s has been classified!'%path)
            print('\n')
            continue
        data = pd.read_excel(path,header=None)
        nums = data.shape[0]
        #save features and class label of the neuron
        print('get neuron features and neuron classes...')
        neuron_features = []
        neuron_classes = []
        for i in range(0,nums,22):
            neuron_info = data.loc[i,0]
            neuron_info = neuron_info.split('\\')
            #default class
            neuron_class = 1
            for j in range(0,len(neuron_info)):
                if neuron_info[j] == 'train_dataset':
                    neuron_class = int(neuron_info[j+1])
                    break
            neuron_feature = data.loc[range(i,i+22),2].values
            neuron_features.append(neuron_feature)
            neuron_classes.append(neuron_class)
        neuron_classes_set = set(neuron_classes)
        if len(neuron_classes_set) == 1:
            print('only one neuron class,can\'t do classification!')
            print('\n')
            continue
        #Z-score normalization
        print('normalization...')
        ss = StandardScaler()
        neuron_features = ss.fit_transform(neuron_features)
        #RFECV to choose the most important features use the whole data
        print('rfecv to get optimal features...')
        svc = SVC(kernel = 'linear')
        from sklearn.model_selection import StratifiedKFold
        #test the best value of StratifiedKFold(value)
        rfecv = RFECV(estimator=svc,step=1,cv = StratifiedKFold(k_fold),scoring = 'accuracy')
        try:
            rfecv.fit(neuron_features,neuron_classes)
        except Exception as e:
            print('file %s has some problems!'%path)
            print(e)
            print('\n')
            problem_pathlist.append(path)
            continue
        #1 represent is the features be chosen
        refecv_ranking = rfecv.ranking_
        remain_feature = []
        for i in range(0,len(refecv_ranking)):
            if refecv_ranking[i] == 1:
                remain_feature.append(i)
        #select the remain_feature on whole data
        neuron_features = np.array(neuron_features)
        neuron_features = neuron_features[:,remain_feature]
        print("Optimal number of features : %d" % rfecv.n_features_)
        print("selected features : %s" % remain_feature)
        #split train_set and test_set
        print('train_test_split...')
        #这个是随机分割，需要避免分割后的train_y和test_y的类别数量和原始数据集总类别数量不一致
        try:
            train_X,test_X,train_y,test_y = train_test_split(neuron_features,neuron_classes,test_size=0.4,stratify=neuron_classes)
        except Exception as e:
            #数据集中某些类别只有一个神经元的特殊情况
            #我考虑人工处理，这样比写算法快，选择直接跳过不分类，人工处理后再次运行。
            #人工处理是直接删除那个神经元的数据
            print('Some neuron class only have one neuron in %s'%path)
            problem_pathlist.append(path)
            continue
        train_y_set = set(train_y)
        test_y_set = set(test_y)
        #SVC classify
        print('SVC classify...')
        model = SVC(kernel='linear')
        model.fit(train_X,train_y)
        #SVC predict
        print('predict...')
        predict_y = model.predict(test_X)
        F_measure = metrics.f1_score(test_y,predict_y,average='weighted')
        print('the F_measure of %s is '%(path),F_measure)
        from sklearn.metrics import classification_report
        #create target_names
        #类别并不一定都是5类，需要动态变化，只要处理非decompose_by_depth和rawdata的数据
        target_names = []
        for i in range(1,class_nums+1):
                if i in neuron_classes_set:
                    target_names.append(str(i))
        #一个潜在坑是，predict_y可能并不会覆盖所有的class，导致其class的数量少于test_y，从而报错 已解决
        #另一个潜在坑是，分割后的test_y class的数量少于训练集类别数量 见上
        classify_result = classification_report(test_y,predict_y,target_names=target_names,output_dict=True,zero_division=0)
        #save the f-measure
        print('save the f-measure result...')
        frame = pd.DataFrame(classify_result,index=None,columns=None)
        frame.to_excel(save_path)
        print('%s has been classified!'% path)
        print('\n')
    print('finish classify!')
    print('These file have some problems:')
    for i in problem_pathlist:
        print(i)
    print('Total classify time: ',time.time() - start_time,' seconds')
    print('\n')

#classify进化版，采用交叉验证方法，去除数据差异性带来的影响
def classify2(data_folder,class_nums,k_fold):
    from sklearn.model_selection import cross_validate
    from sklearn.metrics import classification_report
    start_time = time.time()
    print('start to classify...')
    data_folder = os.path.join(data_folder,'measure_result')
    dir_name = os.listdir(data_folder)
    #path of all output.xlsx
    path_list = []
    #存储有问题的数据位置，人工更改
    problem_pathlist = []
    for i in dir_name:
        if i == 'rawdata':
            path_list.append(os.path.join(data_folder,i,'output.xlsx'))
        elif i in ['decompose_by_depth','decompose_by_height','decompose_by_slice']:
            for j in os.listdir(os.path.join(data_folder,i)):
                path_list.append(os.path.join(data_folder,i,j,'output.xlsx'))
    for path in path_list:
        save_path = os.path.join(os.path.dirname(path),'classify_result.xlsx')
        if os.path.exists(save_path):
            print('file %s has been classified!'%path)
            print('\n')
            continue
        data = pd.read_excel(path,header=None)
        data = data.dropna()
        data = data.reset_index(drop=True)
        nums = data.shape[0]
        #save features and class label of the neuron
        print('get neuron features and neuron classes...')
        neuron_features = []
        neuron_classes = []
        for i in range(0,nums,22):
            neuron_info = data.loc[i,0]
            neuron_info = neuron_info.split('\\')
            #default class
            neuron_class = 1
            for j in range(0,len(neuron_info)):
                if neuron_info[j] == 'train_dataset':
                    neuron_class = int(neuron_info[j+1])
                    break
            neuron_feature = data.loc[range(i,i+22),2].values
            neuron_features.append(neuron_feature)
            neuron_classes.append(neuron_class)
        #todo:有个潜在问题是neuron_classes_set的顺序一定是1-10正序吗？？？
        #解答：一定是正序，set自动排序了。
        neuron_classes_set = set(neuron_classes)
        if len(neuron_classes_set) == 1:
            print('only one neuron class,can\'t do classification!')
            problem_pathlist.append(path)
            print('\n')
            continue
        #Z-score normalization
        print('normalization...')
        ss = StandardScaler()
        neuron_features = ss.fit_transform(neuron_features)
        #RFECV to choose the most important features use the whole data
        print('rfecv to get optimal features...')
        svc = SVC(kernel = 'linear')
        from sklearn.model_selection import StratifiedKFold
        #test the best value of StratifiedKFold(value)
        rfecv = RFECV(estimator=svc,step=1,cv = StratifiedKFold(5),scoring = 'accuracy')
        try:
            rfecv.fit(neuron_features,neuron_classes)
        except Exception as e:
            print('file %s has some problems!'%path)
            print(e)
            print('\n')
            problem_pathlist.append(path)
            continue
        #1 represent is the features be chosen
        refecv_ranking = rfecv.ranking_
        remain_feature = []
        for i in range(0,len(refecv_ranking)):
            if refecv_ranking[i] == 1:
                remain_feature.append(i)
        #select the remain_feature on whole data
        neuron_features = np.array(neuron_features)
        neuron_classes = np.array(neuron_classes)
        neuron_features = neuron_features[:,remain_feature]
        print("Optimal number of features : %d" % rfecv.n_features_)
        print("selected features : %s" % remain_feature)
        #k_fold cross validation
        print('Do %d-fold cross validation...'%(k_fold))
        skf = StratifiedKFold(n_splits = k_fold)
        j = 0
        data = {}
        try:
            for train_index,test_index in skf.split(neuron_features,neuron_classes):
                train_X,test_X,train_y,test_y = neuron_features[train_index],neuron_features[test_index],neuron_classes[train_index],neuron_classes[test_index]
                print('done %d th cross validation!'%(j+1))
                #SVC classify
                print('do %d th SVC classify...'%(j+1))
                model = SVC(kernel='linear')
                model.fit(train_X,train_y)
                #SVC predict
                print('predict...')
                predict_y = model.predict(test_X)
                #create target_names
                #类别并不一定都是全部类，需要动态变化，只要处理非decompose_by_depth和rawdata的数据
                target_names = []
                for i in range(1,class_nums+1):
                        if i in neuron_classes_set:
                            target_names.append(str(i))
                #一个潜在坑是，predict_y可能并不会覆盖所有的class，导致其class的数量少于test_y，从而报错 已解决
                #另一个潜在坑是，分割后的test_y class的数量少于训练集类别数量 见上
                classify_result = classification_report(test_y,predict_y,target_names=target_names,output_dict=True,zero_division=0)
                data[j] = pd.DataFrame(classify_result,index=None,columns=None)
                j = j+1
            #k-fold次验证取平均值
            result_frame = pd.Panel(data).mean(axis=0)
            result_frame.to_excel(save_path)
            print('%s has been classified!'% path)
            print('\n')
        except Exception as e:
            print('k_fold = %d cause error in file %s'%(k_fold,path))
            print(e)
            print('\n')
            problem_pathlist.append(path)
            continue
    print('finish classify!')
    print('These file have some problems:')
    for i in problem_pathlist:
        print(i)
    print('Total classify time: ',time.time() - start_time,' seconds')
    print('\n')


#PCA to find which component influence the classification most
#the data is output.xlsx of raw data
def pca(data_folder,n_components):
    save_path = os.path.join(data_folder,'measure_result','rawdata','pca.xlsx')
    if os.path.exists(save_path):
        print('pca has been caculated!')
    else:
        from sklearn.decomposition import PCA
        start_time = time.time()
        print('start to pca...')
        path  = os.path.join(data_folder,'measure_result','rawdata','output.xlsx')
        #以下可以写成一个独立的方法，功能是从现有output.xlsx中提取出特征和label数据
        data = pd.read_excel(path,header=None)
        nums = data.shape[0]
        neuron_features = []
        neuron_classes = []
        for i in range(0,nums,22):
            neuron_info = data.loc[i,0]
            neuron_info = neuron_info.split('\\')
            #default class
            neuron_class = 1
            for j in range(0,len(neuron_info)):
                    if neuron_info[j] == 'train_dataset':
                        neuron_class = int(neuron_info[j+1])
                        break
            neuron_feature = data.loc[range(i,i+22),2].values
            neuron_features.append(neuron_feature)
            neuron_classes.append(neuron_class)
        #normalization
        ss = StandardScaler()
        neuron_features = ss.fit_transform(neuron_features)
        pca = PCA(n_components = n_components)
        pca.fit(neuron_features)
        print('pca ratio: %s'%pca.explained_variance_ratio_)
        print('pca components: %s'%pca.components_)
        #save the result
        #pca.componets_和pca.explained_variance_ratio_是一一对应关系吗？？如果出错，原因就在这
        pca_result = np.insert(pca.components_.T,0,values=pca.explained_variance_ratio_,axis=0)
        pca_frame = pd.DataFrame(pca_result,
                                 index=['Variance explanation','Soma Surface','N_stems','N_bifs','N_branch','Width','Height',
                                        'Depth','Diameter','Length','Surface','Volume','EucDistance','PathDistance','Branch Order',
                                        'Branch pathlength','Contraction','Fragmentation','Partition asymmetry','Rall Power','Bif ampl local','Bif ampl remote','Fractal Dim'],
                                 columns=['PC1','PC2','PC3'])
        pca_frame.to_excel(save_path)
        print('save pca result at %s'%save_path)
        print('finish pca!')
        print('Total pca time: ',time.time() - start_time,' seconds')
    
#展示分类结果的函数
def show_result(data_folder,class_num,max_level,slice_max_depth,slice_max_size):
    data_folder = os.path.join(data_folder,'measure_result')
    #experiment 1,whole neurons
    #可根据类别数自动绘制
    #读取rawdata的classify_result
    fig1 = plt.figure()
    result_path = os.path.join(data_folder,'rawdata','classify_result.xlsx')
    data = pd.read_excel(result_path,header=None)
    #F_measure的最后一个值是所有类别的平均值
    F_measure = list(data.loc[1,range(1,class_num+1)].values)
    F_measure.append(data.loc[1,class_num+2])
    #横坐标
    label_list = []
    for i in range(1,class_num+1):
        label_list.append('C'+str(i))
    label_list.append('Avg')
     #每个类别曲线对应的颜色，最多考虑10个类
    #先绘制没有normalized的F-measure关系图
    color_list = ['blue','orange','green','red','purple','brown','pink','gray','olive','cyan','black']
    plt.bar(range(1,class_num+2),F_measure,color=color_list,tick_label = label_list)
    plt.ylabel('F-measure')
    save_path = os.path.join(data_folder,'rawdata','classify_result.png')
    plt.savefig(save_path)
    plt.show()
    print('experiment 1 result has been plotted!')
    print('experiment 1 result has been saved in path %s'%save_path)

    #experiment 2,decomposition by depth
    #绘制F-Measure和Depth曲线图
    fig2  = plt.figure()
    depth_list = list(range(0,max_level+1))
    #每行对应一个类别
    #每列对应一个depth_level
    F_measure_list = []
    #F_measure的每个depth_level的平均值
    F_measure_ave = []
    for i in range(0,max_level+1):
        result_path = os.path.join(data_folder,'decompose_by_depth',str(i),'classify_result.xlsx')
        data = pd.read_excel(result_path,header=None)
        f_measure = list(data.loc[1,range(1,class_num+1)].values)
        F_measure_list.append(f_measure)
        F_measure_ave.append(data.loc[1,class_num+2])
    F_measure_list = list(np.array(F_measure_list).T)
    #每个类别轮流绘制
    plt.title('before normalized')
    for i in range(1,class_num+1):
        plt.plot(depth_list,F_measure_list[i-1],color=color_list[i-1],label='C'+str(i))
    plt.plot(depth_list,F_measure_ave,color='black',label='Avg',linewidth=2,linestyle='--')
    plt.legend()
    plt.xlabel('Depth')
    plt.ylabel('F-measure')
    plt.xticks(np.arange(0,max_level+1,1))
    save_path = os.path.join(data_folder,'decompose_by_depth','before_normalized_result.png')
    plt.savefig(save_path)
    plt.show()
    #再绘制已经normalized的F-measure关系图
    fig3 = plt.figure()
    plt.title('after normalized')
    for i in range(1,class_num+1):
        F_measure_list[i-1] = list(np.array(F_measure_list[i-1])/F_measure[i-1])
        plt.plot(depth_list,F_measure_list[i-1],color=color_list[i-1],label='C'+str(i))
    F_measure_ave = list(np.array(F_measure_ave)/F_measure[class_num])
    plt.plot(depth_list,F_measure_ave,color='black',label='Avg',linewidth=2,linestyle='--')
    plt.legend()
    plt.xlabel('Depth')
    plt.ylabel('F-measure')
    plt.xticks(np.arange(0,max_level+1,1))
    save_path = os.path.join(data_folder,'decompose_by_depth','after_normalized_result.png')
    plt.savefig(save_path)
    plt.show()
    print('experiment 2 result has been plotted!')
    print('experiment 2 result has been saved in path %s'%save_path)

    #experiment 3,decomposition by height
    #绘制F-Measure和Height曲线图
    #某些height可能没有分类结果，即没有classify_result.xlsx文件——处理方式：没有分类结果的height_level F-measure直接计算为0
    #有classify_result的，类别也不一定连续——处理方式：某个height_level下，某类别没有结果的 F-measure直接计算为0
    fig4 = plt.figure()
    height_list = list(range(0,max_level+1))
    F_measure_list = np.zeros([class_num,max_level+1])
    F_measure_ave = np.zeros(max_level+1)
    for i in range(0,max_level+1):
        result_path = os.path.join(data_folder,'decompose_by_height',str(i),'classify_result.xlsx')
        if os.path.exists(result_path):
            data = pd.read_excel(result_path,header=None)
            temp = data.loc[0,:]
            #找出最后一个类别所在列
            for idx in range(0,len(temp)):
                if temp[idx]=='accuracy':
                    break
            idx = idx - 1
            for j in range(1,idx+1):
                F_measure_list[int(data.loc[0][j])-1][i] = data.loc[1][j]
            F_measure_ave[i] = data.loc[1][idx+2]
    color_list = ['blue','orange','green','red','purple','brown','pink','gray','olive','cyan']
    plt.title('before normalized')
    for i in range(1,class_num+1):
        plt.plot(depth_list,F_measure_list[i-1],color=color_list[i-1],label='C'+str(i))
    plt.plot(depth_list,F_measure_ave,color='black',label='Avg',linewidth=2,linestyle='--')
    plt.legend()
    plt.xlabel('Height')
    plt.ylabel('F-measure')
    plt.xticks(np.arange(0,max_level+1,1))
    save_path = os.path.join(data_folder,'decompose_by_height','before_normalized_result.png')
    plt.savefig(save_path)
    plt.show()
    #再绘制已经normalized的F-measure关系图
    fig5 = plt.figure()
    plt.title('after normalized')
    for i in range(1,class_num+1):
        F_measure_list[i-1] = list(np.array(F_measure_list[i-1])/F_measure[i-1])
        plt.plot(depth_list,F_measure_list[i-1],color=color_list[i-1],label='C'+str(i))
    F_measure_ave = list(np.array(F_measure_ave)/F_measure[class_num])
    plt.plot(depth_list,F_measure_ave,color='black',label='Avg',linewidth=2,linestyle='--')
    plt.legend()
    plt.xlabel('Height')
    plt.ylabel('F-measure')
    plt.xticks(np.arange(0,max_level+1,1))
    save_path = os.path.join(data_folder,'decompose_by_height','after_normalized_result.png')
    plt.savefig(save_path)
    plt.show()
    print('experiment 3 result has been plotted!')
    print('experiment 3 result has been saved in path %s'%save_path)

    #experiment 4,decomposition by slices
    #绘制slice_depth和slice_size热力图，只取每个level的平均值
    #对应文件夹下没有classify_result.xlsx:F-measure设置为0
    #有的classify_result.xlsx可能不包含所有类的结果：不要默认classify_result.xlsx一定有所有类别的F-measure
    #每行代表对应slice_depth的值，每列代表对应slice_size的值，每个元素代表对应的F-measure
    fig6 = plt.figure()
    F_measure_list = np.zeros([slice_max_depth+1,slice_max_size])
    for slice_depth in range(0,slice_max_depth+1):
        for slice_size in range(1,slice_max_size+1):
            result_path = os.path.join(data_folder,'decompose_by_slice','slice_{}_{}'.format(slice_depth,slice_size),'classify_result.xlsx')
            if os.path.exists(result_path):
                data = pd.read_excel(result_path,header=None)
                temp = data.loc[0,:]
                #找出最后一个类别所在列
                for idx in range(0,len(temp)):
                    if temp[idx]=='accuracy':
                        break
                idx = idx - 1
                #这里注意从上到下，为slice_max_depth、slice_max_depth-1,...,2,1,0，为了画出和论文中相同的效果
                F_measure_list[slice_max_depth-slice_depth][slice_size-1]=data.loc[1,idx+2]
    import seaborn as sns
    sns.set()
    #F_measure_list转为DataFrame
    index = list(range(0,slice_max_depth+1))
    index.reverse()
    columns = list(range(1,slice_max_size+1))
    F_measure_list = pd.DataFrame(F_measure_list,index=index,columns=columns)
    print(F_measure_list)
    #F_measure_list = F_measure_list.pivot('Slice depth','Slice size','F-measure')
    sns.heatmap(F_measure_list,cmap='YlGnBu',annot=True,fmt='.2',cbar_kws={'label': 'F-measure'})
    sns.despine(right=True)
    plt.xlabel('Slice size')
    plt.ylabel('Slice depth')
    save_path = os.path.join(data_folder,'decompose_by_slice','slice_heatmap.png')
    plt.savefig(save_path)
    plt.show()
    print('experiment 4 result has been plotted!')
    print('experiment  result has been saved in path %s'%save_path)

    #experiment 5,Decomposition by Slice in Rat and Mouse Species


    #experiment 6,Decomposition by Slice in Neocortical Region




if __name__=='__main__':
    root_path = 'D:\\Study\\The Road to AI\\AIRoad\\neuro_data\\train_dataset'
    class_num = 10
    #max_depth  and max_height
    #for this dataset
    #todo:记得改回来
    #max_level = get_maxlevel(root_path,class_num)
    max_level = 35
    slice_max_depth = 10
    slice_max_size = max_level
    #decompose(root_path,class_num,max_level,slice_max_depth,slice_max_size)
    #after decompose have to wait the result of L-Measure
    #save folder of the output.xlsx
    save_folder = os.path.join(root_path,'measure_data')
    #this file save raw data and all decomposed neuron mophological data
    #use L-Measure to measure all neurons under root_path,and save it to path below
    #eg:output1.xlsx save the data of class 1(composed of decompose by depth,height and rawdata).
    path =  'D:\\Study\\The Road to AI\\AIRoad\\neuro_data\\train_dataset\\totaldata_measure\\output{}.xlsx'
    if os.path.exists(path.format(1)):
        #preprocess_data2(path,root_path,class_num)
        #k folds validation
        #论文里是10，需要测试哪个合适
        k_fold = 10
        #classify2(root_path,class_num,k_fold)
        #pca(root_path,3)
        show_result(root_path,class_num,max_level,slice_max_depth,slice_max_size)
    else:
        print('please use L-Measure to measure all neurons under root_path,and save it to path:%s'%path)


